import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import os, sys


def function_main_Plotting_Division_Win_Pctg(selected_Column):

    # Open a file
    path = "./NBA_Data/"
    dirs = os.listdir(path)

    # ========================================
    # Define parameters
    Rank = 1
    Team = 2
    Wins = 3
    Loss = 4
    Win_Pctg = 5
    Win_Diff = 6
    Home_Result = 7
    Road_Result = 8
    Div_Result = 9
    Conf_Result = 10
    Pt_Gain = 11
    Pt_Loss = 12
    Pt_Diff = 13
    WL_Streak = 14

    # Define parameters
    # East Teams
    Rank_all_TRR = 0
    Rank_all_MWB = 0
    Rank_all_IDP = 0
    Rank_all_P76 = 0
    Rank_all_DPT = 0
    Rank_all_OLM = 0
    Rank_all_BOC = 0
    Rank_all_CLH = 0
    Rank_all_BRN = 0
    Rank_all_MIH = 0
    Rank_all_WSW = 0
    Rank_all_CCB = 0
    Rank_all_NYN = 0
    Rank_all_ATH = 0
    Rank_all_CLC = 0

    # West Teams
    Rank_all_PTB = 0
    Rank_all_MGR = 0
    Rank_all_LAC = 0
    Rank_all_GSW = 0
    Rank_all_OKC = 0
    Rank_all_DVN = 0
    Rank_all_NOP = 0
    Rank_all_LAL = 0
    Rank_all_HOR = 0
    Rank_all_SAK = 0
    Rank_all_SAS = 0
    Rank_all_UTJ = 0
    Rank_all_DAM = 0
    Rank_all_MTW = 0
    Rank_all_PHS = 0

    File_Date_all = 0

    # ========================================
    # Sweep all files in the folder
    for filename in dirs:
        filepath = './NBA_Data/' + filename
        wb = openpyxl.load_workbook(filepath)
        sheet1 = wb['Sheet1']
        sheet2 = wb['Sheet2']

        # ------------------------------------------
        # Ranking
        # ------------------------------------------
        # East Conference Ranking
        for i in range(1, 16):
            if ((sheet1.cell(row=i + 1, column=Team).value) == '猛龙'):
                Rank_val_TRR = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_TRR = Rank_val_TRR.split('-')
                Rank_val_TRR_W = float(Rank_val_TRR[0])
                Rank_val_TRR_L = float(Rank_val_TRR[1])
                if ((Rank_val_TRR_W == 0) and (Rank_val_TRR_L == 0)):
                    Rank_val_TRR = 0
                else:
                    Rank_val_TRR = 100 * Rank_val_TRR_W / (Rank_val_TRR_W + Rank_val_TRR_L)
                Rank_all_TRR = np.append(Rank_all_TRR, Rank_val_TRR)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '雄鹿'):
                Rank_val_MWB = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_MWB = Rank_val_MWB.split('-')
                Rank_val_MWB_W = float(Rank_val_MWB[0])
                Rank_val_MWB_L = float(Rank_val_MWB[1])
                if ((Rank_val_MWB_W == 0) and (Rank_val_MWB_L == 0)):
                    Rank_val_MWB = 0
                else:
                    Rank_val_MWB = 100 * Rank_val_MWB_W / (Rank_val_MWB_W + Rank_val_MWB_L)
                Rank_all_MWB = np.append(Rank_all_MWB, Rank_val_MWB)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '步行者'):
                Rank_val_IDP = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_IDP = Rank_val_IDP.split('-')
                Rank_val_IDP_W = float(Rank_val_IDP[0])
                Rank_val_IDP_L = float(Rank_val_IDP[1])
                if ((Rank_val_IDP_W == 0) and (Rank_val_IDP_L == 0)):
                    Rank_val_IDP = 0
                else:
                    Rank_val_IDP = 100 * Rank_val_IDP_W / (Rank_val_IDP_W + Rank_val_IDP_L)
                Rank_all_IDP = np.append(Rank_all_IDP, Rank_val_IDP)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '76人'):
                Rank_val_P76 = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_P76 = Rank_val_P76.split('-')
                Rank_val_P76_W = float(Rank_val_P76[0])
                Rank_val_P76_L = float(Rank_val_P76[1])
                if ((Rank_val_P76_W == 0) and (Rank_val_P76_L == 0)):
                    Rank_val_P76 = 0
                else:
                    Rank_val_P76 = 100 * Rank_val_P76_W / (Rank_val_P76_W + Rank_val_P76_L)
                Rank_all_P76 = np.append(Rank_all_P76, Rank_val_P76)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '活塞'):
                Rank_val_DPT = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_DPT = Rank_val_DPT.split('-')
                Rank_val_DPT_W = float(Rank_val_DPT[0])
                Rank_val_DPT_L = float(Rank_val_DPT[1])
                if ((Rank_val_DPT_W == 0) and (Rank_val_DPT_L == 0)):
                    Rank_val_DPT = 0
                else:
                    Rank_val_DPT = 100 * Rank_val_DPT_W / (Rank_val_DPT_W + Rank_val_DPT_L)
                Rank_all_DPT = np.append(Rank_all_DPT, Rank_val_DPT)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '魔术'):
                Rank_val_OLM = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_OLM = Rank_val_OLM.split('-')
                Rank_val_OLM_W = float(Rank_val_OLM[0])
                Rank_val_OLM_L = float(Rank_val_OLM[1])
                if ((Rank_val_OLM_W == 0) and (Rank_val_OLM_L == 0)):
                    Rank_val_OLM = 0
                else:
                    Rank_val_OLM = 100 * Rank_val_OLM_W / (Rank_val_OLM_W + Rank_val_OLM_L)
                Rank_all_OLM = np.append(Rank_all_OLM, Rank_val_OLM)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '凯尔特人'):
                Rank_val_BOC = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_BOC = Rank_val_BOC.split('-')
                Rank_val_BOC_W = float(Rank_val_BOC[0])
                Rank_val_BOC_L = float(Rank_val_BOC[1])
                if ((Rank_val_BOC_W == 0) and (Rank_val_BOC_L == 0)):
                    Rank_val_BOC = 0
                else:
                    Rank_val_BOC = 100 * Rank_val_BOC_W / (Rank_val_BOC_W + Rank_val_BOC_L)
                Rank_all_BOC = np.append(Rank_all_BOC, Rank_val_BOC)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '黄蜂'):
                Rank_val_CLH = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_CLH = Rank_val_CLH.split('-')
                Rank_val_CLH_W = float(Rank_val_CLH[0])
                Rank_val_CLH_L = float(Rank_val_CLH[1])
                if ((Rank_val_CLH_W == 0) and (Rank_val_CLH_L == 0)):
                    Rank_val_CLH = 0
                else:
                    Rank_val_CLH = 100 * Rank_val_CLH_W / (Rank_val_CLH_W + Rank_val_CLH_L)
                Rank_all_CLH = np.append(Rank_all_CLH, Rank_val_CLH)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '篮网'):
                Rank_val_BRN = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_BRN = Rank_val_BRN.split('-')
                Rank_val_BRN_W = float(Rank_val_BRN[0])
                Rank_val_BRN_L = float(Rank_val_BRN[1])
                if ((Rank_val_BRN_W == 0) and (Rank_val_BRN_L == 0)):
                    Rank_val_BRN = 0
                else:
                    Rank_val_BRN = 100 * Rank_val_BRN_W / (Rank_val_BRN_W + Rank_val_BRN_L)
                Rank_all_BRN = np.append(Rank_all_BRN, Rank_val_BRN)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '热火'):
                Rank_val_MIH = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_MIH = Rank_val_MIH.split('-')
                Rank_val_MIH_W = float(Rank_val_MIH[0])
                Rank_val_MIH_L = float(Rank_val_MIH[1])
                if ((Rank_val_MIH_W == 0) and (Rank_val_MIH_L == 0)):
                    Rank_val_MIH = 0
                else:
                    Rank_val_MIH = 100 * Rank_val_MIH_W / (Rank_val_MIH_W + Rank_val_MIH_L)
                Rank_all_MIH = np.append(Rank_all_MIH, Rank_val_MIH)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '奇才'):
                Rank_val_WSW = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_WSW = Rank_val_WSW.split('-')
                Rank_val_WSW_W = float(Rank_val_WSW[0])
                Rank_val_WSW_L = float(Rank_val_WSW[1])
                if ((Rank_val_WSW_W == 0) and (Rank_val_WSW_L == 0)):
                    Rank_val_WSW = 0
                else:
                    Rank_val_WSW = 100 * Rank_val_WSW_W / (Rank_val_WSW_W + Rank_val_WSW_L)
                Rank_all_WSW = np.append(Rank_all_WSW, Rank_val_WSW)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '公牛'):
                Rank_val_CCB = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_CCB = Rank_val_CCB.split('-')
                Rank_val_CCB_W = float(Rank_val_CCB[0])
                Rank_val_CCB_L = float(Rank_val_CCB[1])
                if ((Rank_val_CCB_W == 0) and (Rank_val_CCB_L == 0)):
                    Rank_val_CCB = 0
                else:
                    Rank_val_CCB = 100 * Rank_val_CCB_W / (Rank_val_CCB_W + Rank_val_CCB_L)
                Rank_all_CCB = np.append(Rank_all_CCB, Rank_val_CCB)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '尼克斯'):
                Rank_val_NYN = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_NYN = Rank_val_NYN.split('-')
                Rank_val_NYN_W = float(Rank_val_NYN[0])
                Rank_val_NYN_L = float(Rank_val_NYN[1])
                if ((Rank_val_NYN_W == 0) and (Rank_val_NYN_L == 0)):
                    Rank_val_NYN = 0
                else:
                    Rank_val_NYN = 100 * Rank_val_NYN_W / (Rank_val_NYN_W + Rank_val_NYN_L)
                Rank_all_NYN = np.append(Rank_all_NYN, Rank_val_NYN)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '老鹰'):
                Rank_val_ATH = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_ATH = Rank_val_ATH.split('-')
                Rank_val_ATH_W = float(Rank_val_ATH[0])
                Rank_val_ATH_L = float(Rank_val_ATH[1])
                if ((Rank_val_ATH_W == 0) and (Rank_val_ATH_L == 0)):
                    Rank_val_ATH = 0
                else:
                    Rank_val_ATH = 100 * Rank_val_ATH_W / (Rank_val_ATH_W + Rank_val_ATH_L)
                Rank_all_ATH = np.append(Rank_all_ATH, Rank_val_ATH)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '骑士'):
                Rank_val_CLC = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_CLC = Rank_val_CLC.split('-')
                Rank_val_CLC_W = float(Rank_val_CLC[0])
                Rank_val_CLC_L = float(Rank_val_CLC[1])
                if ((Rank_val_CLC_W == 0) and (Rank_val_CLC_L == 0)):
                    Rank_val_CLC = 0
                else:
                    Rank_val_CLC = 100 * Rank_val_CLC_W / (Rank_val_CLC_W + Rank_val_CLC_L)
                Rank_all_CLC = np.append(Rank_all_CLC, Rank_val_CLC)

            # West Conference Ranking
        for i in range(1, 16):
            if ((sheet2.cell(row=i + 1, column=Team).value) == '开拓者'):
                Rank_val_PTB = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_PTB = Rank_val_PTB.split('-')
                Rank_val_PTB_W = float(Rank_val_PTB[0])
                Rank_val_PTB_L = float(Rank_val_PTB[1])
                if ((Rank_val_PTB_W == 0) and (Rank_val_PTB_L == 0)):
                    Rank_val_PTB = 0
                else:
                    Rank_val_PTB = 100 * Rank_val_PTB_W / (Rank_val_PTB_W + Rank_val_PTB_L)
                Rank_all_PTB = np.append(Rank_all_PTB, Rank_val_PTB)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '灰熊'):
                Rank_val_MGR = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_MGR = Rank_val_MGR.split('-')
                Rank_val_MGR_W = float(Rank_val_MGR[0])
                Rank_val_MGR_L = float(Rank_val_MGR[1])
                if ((Rank_val_MGR_W == 0) and (Rank_val_MGR_L == 0)):
                    Rank_val_MGR = 0
                else:
                    Rank_val_MGR = 100 * Rank_val_MGR_W / (Rank_val_MGR_W + Rank_val_MGR_L)
                Rank_all_MGR = np.append(Rank_all_MGR, Rank_val_MGR)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '快船'):
                Rank_val_LAC = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_LAC = Rank_val_LAC.split('-')
                Rank_val_LAC_W = float(Rank_val_LAC[0])
                Rank_val_LAC_L = float(Rank_val_LAC[1])
                if ((Rank_val_LAC_W == 0) and (Rank_val_LAC_L == 0)):
                    Rank_val_LAC = 0
                else:
                    Rank_val_LAC = 100 * Rank_val_LAC_W / (Rank_val_LAC_W + Rank_val_LAC_L)
                Rank_all_LAC = np.append(Rank_all_LAC, Rank_val_LAC)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '勇士'):
                Rank_val_GSW = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_GSW = Rank_val_GSW.split('-')
                Rank_val_GSW_W = float(Rank_val_GSW[0])
                Rank_val_GSW_L = float(Rank_val_GSW[1])
                if ((Rank_val_GSW_W == 0) and (Rank_val_GSW_L == 0)):
                    Rank_val_GSW = 0
                else:
                    Rank_val_GSW = 100 * Rank_val_GSW_W / (Rank_val_GSW_W + Rank_val_GSW_L)
                Rank_all_GSW = np.append(Rank_all_GSW, Rank_val_GSW)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '雷霆'):
                Rank_val_OKC = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_OKC = Rank_val_OKC.split('-')
                Rank_val_OKC_W = float(Rank_val_OKC[0])
                Rank_val_OKC_L = float(Rank_val_OKC[1])
                if ((Rank_val_OKC_W == 0) and (Rank_val_OKC_L == 0)):
                    Rank_val_OKC = 0
                else:
                    Rank_val_OKC = 100 * Rank_val_OKC_W / (Rank_val_OKC_W + Rank_val_OKC_L)
                Rank_all_OKC = np.append(Rank_all_OKC, Rank_val_OKC)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '掘金'):
                Rank_val_DVN = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_DVN = Rank_val_DVN.split('-')
                Rank_val_DVN_W = float(Rank_val_DVN[0])
                Rank_val_DVN_L = float(Rank_val_DVN[1])
                if ((Rank_val_DVN_W == 0) and (Rank_val_DVN_L == 0)):
                    Rank_val_DVN = 0
                else:
                    Rank_val_DVN = 100 * Rank_val_DVN_W / (Rank_val_DVN_W + Rank_val_DVN_L)
                Rank_all_DVN = np.append(Rank_all_DVN, Rank_val_DVN)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '鹈鹕'):
                Rank_val_NOP = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_NOP = Rank_val_NOP.split('-')
                Rank_val_NOP_W = float(Rank_val_NOP[0])
                Rank_val_NOP_L = float(Rank_val_NOP[1])
                if ((Rank_val_NOP_W == 0) and (Rank_val_NOP_L == 0)):
                    Rank_val_NOP = 0
                else:
                    Rank_val_NOP = 100 * Rank_val_NOP_W / (Rank_val_NOP_W + Rank_val_NOP_L)
                Rank_all_NOP = np.append(Rank_all_NOP, Rank_val_NOP)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '湖人'):
                Rank_val_LAL = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_LAL = Rank_val_LAL.split('-')
                Rank_val_LAL_W = float(Rank_val_LAL[0])
                Rank_val_LAL_L = float(Rank_val_LAL[1])
                if ((Rank_val_LAL_W == 0) and (Rank_val_LAL_L == 0)):
                    Rank_val_LAL = 0
                else:
                    Rank_val_LAL = 100 * Rank_val_LAL_W / (Rank_val_LAL_W + Rank_val_LAL_L)
                Rank_all_LAL = np.append(Rank_all_LAL, Rank_val_LAL)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '火箭'):
                Rank_val_HOR = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_HOR = Rank_val_HOR.split('-')
                Rank_val_HOR_W = float(Rank_val_HOR[0])
                Rank_val_HOR_L = float(Rank_val_HOR[1])
                if ((Rank_val_HOR_W == 0) and (Rank_val_HOR_L == 0)):
                    Rank_val_HOR = 0
                else:
                    Rank_val_HOR = 100 * Rank_val_HOR_W / (Rank_val_HOR_W + Rank_val_HOR_L)
                Rank_all_HOR = np.append(Rank_all_HOR, Rank_val_HOR)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '国王'):
                Rank_val_SAK = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_SAK = Rank_val_SAK.split('-')
                Rank_val_SAK_W = float(Rank_val_SAK[0])
                Rank_val_SAK_L = float(Rank_val_SAK[1])
                if ((Rank_val_SAK_W == 0) and (Rank_val_SAK_L == 0)):
                    Rank_val_SAK = 0
                else:
                    Rank_val_SAK = 100 * Rank_val_SAK_W / (Rank_val_SAK_W + Rank_val_SAK_L)
                Rank_all_SAK = np.append(Rank_all_SAK, Rank_val_SAK)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '马刺'):
                Rank_val_SAS = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_SAS = Rank_val_SAS.split('-')
                Rank_val_SAS_W = float(Rank_val_SAS[0])
                Rank_val_SAS_L = float(Rank_val_SAS[1])
                if ((Rank_val_SAS_W == 0) and (Rank_val_SAS_L == 0)):
                    Rank_val_SAS = 0
                else:
                    Rank_val_SAS = 100 * Rank_val_SAS_W / (Rank_val_SAS_W + Rank_val_SAS_L)
                Rank_all_SAS = np.append(Rank_all_SAS, Rank_val_SAS)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '爵士'):
                Rank_val_UTJ = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_UTJ = Rank_val_UTJ.split('-')
                Rank_val_UTJ_W = float(Rank_val_UTJ[0])
                Rank_val_UTJ_L = float(Rank_val_UTJ[1])
                if ((Rank_val_UTJ_W == 0) and (Rank_val_UTJ_L == 0)):
                    Rank_val_UTJ = 0
                else:
                    Rank_val_UTJ = 100 * Rank_val_UTJ_W / (Rank_val_UTJ_W + Rank_val_UTJ_L)
                Rank_all_UTJ = np.append(Rank_all_UTJ, Rank_val_UTJ)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '独行侠'):
                Rank_val_DAM = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_DAM = Rank_val_DAM.split('-')
                Rank_val_DAM_W = float(Rank_val_DAM[0])
                Rank_val_DAM_L = float(Rank_val_DAM[1])
                if ((Rank_val_DAM_W == 0) and (Rank_val_DAM_L == 0)):
                    Rank_val_DAM = 0
                else:
                    Rank_val_DAM = 100 * Rank_val_DAM_W / (Rank_val_DAM_W + Rank_val_DAM_L)
                Rank_all_DAM = np.append(Rank_all_DAM, Rank_val_DAM)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '森林狼'):
                Rank_val_MTW = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_MTW = Rank_val_MTW.split('-')
                Rank_val_MTW_W = float(Rank_val_MTW[0])
                Rank_val_MTW_L = float(Rank_val_MTW[1])
                if ((Rank_val_MTW_W == 0) and (Rank_val_MTW_L == 0)):
                    Rank_val_MTW = 0
                else:
                    Rank_val_MTW = 100 * Rank_val_MTW_W / (Rank_val_MTW_W + Rank_val_MTW_L)
                Rank_all_MTW = np.append(Rank_all_MTW, Rank_val_MTW)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '太阳'):
                Rank_val_PHS = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_PHS = Rank_val_PHS.split('-')
                Rank_val_PHS_W = float(Rank_val_PHS[0])
                Rank_val_PHS_L = float(Rank_val_PHS[1])
                if ((Rank_val_PHS_W == 0) and (Rank_val_PHS_L == 0)):
                    Rank_val_PHS = 0
                else:
                    Rank_val_PHS = 100 * Rank_val_PHS_W / (Rank_val_PHS_W + Rank_val_PHS_L)
                Rank_all_PHS = np.append(Rank_all_PHS, Rank_val_PHS)
        # ------------------------------------------

    # East Team Ranking Figure
    f = plt.figure()
    ax = f.add_subplot(111)
    plt.plot(Rank_all_TRR[1:], 'ro-', label='Toronto Raptors')
    plt.plot(Rank_all_MWB[1:], 'go-', label='Milwaukee Bucks')
    plt.plot(Rank_all_IDP[1:], 'bo-', label='Indiana Pacers')
    plt.plot(Rank_all_P76[1:], 'r*-', label='Philadelphia 76ers')
    plt.plot(Rank_all_DPT[1:], 'g*-', label='Detroit Pistons')
    plt.plot(Rank_all_OLM[1:], 'b*-', label='Orlando Magics')
    plt.plot(Rank_all_BOC[1:], 'rx-', label='Boston Celtics')
    plt.plot(Rank_all_CLH[1:], 'gx-', label='Charlotte Hornets')
    plt.plot(Rank_all_BRN[1:], 'bx-', label='Brooklyn Nets')
    plt.plot(Rank_all_MIH[1:], 'rd-', label='Miami Heats')
    plt.plot(Rank_all_WSW[1:], 'gd-', label='Washington Wizards')
    plt.plot(Rank_all_CCB[1:], 'bd-', label='Chicago Bulls')
    plt.plot(Rank_all_NYN[1:], 'rs-', label='New York Knicks')
    plt.plot(Rank_all_ATH[1:], 'gs-', label='Atlanta Hawks')
    plt.plot(Rank_all_CLC[1:], 'bs-', label='Cleveland Cavaliers')
    plt.title('East Teams Division Winning Percentage History')
    plt.xlabel('Time')
    plt.ylabel('Percentage (%)')
    ax.yaxis.set_ticks_position('both')
    leg = plt.legend(loc='upper left', fancybox=True, fontsize=9)
    leg.get_frame().set_alpha(0.2)
    # for x_tick in range(0, num_Files):
    #     plt.text(x_tick, 110, File_Date_all[x_tick], rotation=45)

    # West Team Ranking Figure
    f = plt.figure()
    ax = f.add_subplot(111)
    plt.plot(Rank_all_PTB[1:], 'c-', label='Portland Trail Blazers')
    plt.plot(Rank_all_MGR[1:], 'yo-', label='Memphis Grizzlies')
    plt.plot(Rank_all_LAC[1:], 'ko-', label='LA Clippers')
    plt.plot(Rank_all_GSW[1:], 'c*-', label='Golden State Warriors')
    plt.plot(Rank_all_OKC[1:], 'y*-', label='Oklahoma City Thunders')
    plt.plot(Rank_all_DVN[1:], 'k*-', label='Denver Nuggets')
    plt.plot(Rank_all_NOP[1:], 'cx-', label='New Orleans Pelicans')
    plt.plot(Rank_all_LAL[1:], 'yx-', label='Los Angeles Lakers')
    plt.plot(Rank_all_HOR[1:], 'kx-', label='Houston Rockets')
    plt.plot(Rank_all_SAK[1:], 'cd-', label='Sacramento Kings')
    plt.plot(Rank_all_SAS[1:], 'yd-', label='San Antonio Spurs')
    plt.plot(Rank_all_UTJ[1:], 'kd-', label='Utah Jazz')
    plt.plot(Rank_all_DAM[1:], 'cs-', label='Dallas Mavericks')
    plt.plot(Rank_all_MTW[1:], 'ys-', label='Minnesota Timberwolves')
    plt.plot(Rank_all_PHS[1:], 'ks-', label='Phoenix Suns')
    plt.title('West Teams Division Winning Percentage History')
    plt.xlabel('Time')
    plt.ylabel('Percentage (%)')
    ax.yaxis.set_ticks_position('both')
    leg = plt.legend(loc='upper left', fancybox=True, fontsize=9)
    leg.get_frame().set_alpha(0.2)
    # for x_tick in range(0, num_Files):
    #     plt.text(x_tick, 110, File_Date_all[x_tick], rotation=45)


    return
