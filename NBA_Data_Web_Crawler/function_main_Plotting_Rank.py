import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import os, sys


def function_main_Plotting_Rank(selected_Column):

    # Open a file
    path = "./NBA_Data/"
    dirs = os.listdir(path)

    # ========================================
    # Define parameters
    Rank = 1
    Team = 2
    Wins = 3
    Loss = 4
    Win_Pctg = 5
    Win_Diff = 6
    Home_Result = 7
    Road_Result = 8
    Div_Result = 9
    Conf_Result = 10
    Pt_Gain = 11
    Pt_Loss = 12
    Pt_Diff = 13
    WL_Streak = 14

    # Define parameters
    # East Teams
    Rank_all_TRR = 0
    Rank_all_MWB = 0
    Rank_all_IDP = 0
    Rank_all_P76 = 0
    Rank_all_DPT = 0
    Rank_all_OLM = 0
    Rank_all_BOC = 0
    Rank_all_CLH = 0
    Rank_all_BRN = 0
    Rank_all_MIH = 0
    Rank_all_WSW = 0
    Rank_all_CCB = 0
    Rank_all_NYN = 0
    Rank_all_ATH = 0
    Rank_all_CLC = 0

    # West Teams
    Rank_all_PTB = 0
    Rank_all_MGR = 0
    Rank_all_LAC = 0
    Rank_all_GSW = 0
    Rank_all_OKC = 0
    Rank_all_DVN = 0
    Rank_all_NOP = 0
    Rank_all_LAL = 0
    Rank_all_HOR = 0
    Rank_all_SAK = 0
    Rank_all_SAS = 0
    Rank_all_UTJ = 0
    Rank_all_DAM = 0
    Rank_all_MTW = 0
    Rank_all_PHS = 0

    File_Date_all = 0

    # ========================================
    # Sweep all files in the folder
    for filename in dirs:
        filepath = './NBA_Data/' + filename
        wb = openpyxl.load_workbook(filepath)
        sheet1 = wb['Sheet1']
        sheet2 = wb['Sheet2']

        File_Date_all = np.append(File_Date_all, filename)

        # ------------------------------------------
        # Ranking
        # ------------------------------------------
        # East Conference Ranking
        for i in range(1, 16):
            if ((sheet1.cell(row=i + 1, column=Team).value) == '猛龙'):
                Rank_val_TRR = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_TRR = float(Rank_val_TRR)
                Rank_all_TRR = np.append(Rank_all_TRR, Rank_val_TRR)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '雄鹿'):
                Rank_val_MWB = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_MWB = float(Rank_val_MWB)
                Rank_all_MWB = np.append(Rank_all_MWB, Rank_val_MWB)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '步行者'):
                Rank_val_IDP = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_IDP = float(Rank_val_IDP)
                Rank_all_IDP = np.append(Rank_all_IDP, Rank_val_IDP)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '76人'):
                Rank_val_P76 = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_P76 = float(Rank_val_P76)
                Rank_all_P76 = np.append(Rank_all_P76, Rank_val_P76)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '活塞'):
                Rank_val_DPT = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_DPT = float(Rank_val_DPT)
                Rank_all_DPT = np.append(Rank_all_DPT, Rank_val_DPT)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '魔术'):
                Rank_val_OLM = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_OLM = float(Rank_val_OLM)
                Rank_all_OLM = np.append(Rank_all_OLM, Rank_val_OLM)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '凯尔特人'):
                Rank_val_BOC = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_BOC = float(Rank_val_BOC)
                Rank_all_BOC = np.append(Rank_all_BOC, Rank_val_BOC)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '黄蜂'):
                Rank_val_CLH = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_CLH = float(Rank_val_CLH)
                Rank_all_CLH = np.append(Rank_all_CLH, Rank_val_CLH)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '篮网'):
                Rank_val_BRN = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_BRN = float(Rank_val_BRN)
                Rank_all_BRN = np.append(Rank_all_BRN, Rank_val_BRN)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '热火'):
                Rank_val_MIH = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_MIH = float(Rank_val_MIH)
                Rank_all_MIH = np.append(Rank_all_MIH, Rank_val_MIH)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '奇才'):
                Rank_val_WSW = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_WSW = float(Rank_val_WSW)
                Rank_all_WSW = np.append(Rank_all_WSW, Rank_val_WSW)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '公牛'):
                Rank_val_CCB = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_CCB = float(Rank_val_CCB)
                Rank_all_CCB = np.append(Rank_all_CCB, Rank_val_CCB)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '尼克斯'):
                Rank_val_NYN = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_NYN = float(Rank_val_NYN)
                Rank_all_NYN = np.append(Rank_all_NYN, Rank_val_NYN)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '老鹰'):
                Rank_val_ATH = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_ATH = float(Rank_val_ATH)
                Rank_all_ATH = np.append(Rank_all_ATH, Rank_val_ATH)

            elif ((sheet1.cell(row=i + 1, column=Team).value) == '骑士'):
                Rank_val_CLC = sheet1.cell(row=i + 1, column=selected_Column).value
                Rank_val_CLC = float(Rank_val_CLC)
                Rank_all_CLC = np.append(Rank_all_CLC, Rank_val_CLC)

        # West Conference Ranking
        for i in range(1, 16):
            if ((sheet2.cell(row=i + 1, column=Team).value) == '开拓者'):
                Rank_val_PTB = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_PTB = float(Rank_val_PTB)
                Rank_all_PTB = np.append(Rank_all_PTB, Rank_val_PTB)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '灰熊'):
                Rank_val_MGR = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_MGR = float(Rank_val_MGR)
                Rank_all_MGR = np.append(Rank_all_MGR, Rank_val_MGR)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '快船'):
                Rank_val_LAC = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_LAC = float(Rank_val_LAC)
                Rank_all_LAC = np.append(Rank_all_LAC, Rank_val_LAC)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '勇士'):
                Rank_val_GSW = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_GSW = float(Rank_val_GSW)
                Rank_all_GSW = np.append(Rank_all_GSW, Rank_val_GSW)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '雷霆'):
                Rank_val_OKC = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_OKC = float(Rank_val_OKC)
                Rank_all_OKC = np.append(Rank_all_OKC, Rank_val_OKC)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '掘金'):
                Rank_val_DVN = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_DVN = float(Rank_val_DVN)
                Rank_all_DVN = np.append(Rank_all_DVN, Rank_val_DVN)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '鹈鹕'):
                Rank_val_NOP = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_NOP = float(Rank_val_NOP)
                Rank_all_NOP = np.append(Rank_all_NOP, Rank_val_NOP)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '湖人'):
                Rank_val_LAL = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_LAL = float(Rank_val_LAL)
                Rank_all_LAL = np.append(Rank_all_LAL, Rank_val_LAL)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '火箭'):
                Rank_val_HOR = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_HOR = float(Rank_val_HOR)
                Rank_all_HOR = np.append(Rank_all_HOR, Rank_val_HOR)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '国王'):
                Rank_val_SAK = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_SAK = float(Rank_val_SAK)
                Rank_all_SAK = np.append(Rank_all_SAK, Rank_val_SAK)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '马刺'):
                Rank_val_SAS = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_SAS = float(Rank_val_SAS)
                Rank_all_SAS = np.append(Rank_all_SAS, Rank_val_SAS)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '爵士'):
                Rank_val_UTJ = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_UTJ = float(Rank_val_UTJ)
                Rank_all_UTJ = np.append(Rank_all_UTJ, Rank_val_UTJ)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '独行侠'):
                Rank_val_DAM = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_DAM = float(Rank_val_DAM)
                Rank_all_DAM = np.append(Rank_all_DAM, Rank_val_DAM)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '森林狼'):
                Rank_val_MTW = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_MTW = float(Rank_val_MTW)
                Rank_all_MTW = np.append(Rank_all_MTW, Rank_val_MTW)

            elif ((sheet2.cell(row=i + 1, column=Team).value) == '太阳'):
                Rank_val_PHS = sheet2.cell(row=i + 1, column=selected_Column).value
                Rank_val_PHS = float(Rank_val_PHS)
                Rank_all_PHS = np.append(Rank_all_PHS, Rank_val_PHS)
        # ------------------------------------------

    File_Date_all = np.transpose(np.matrix(File_Date_all))
    File_Date_all = File_Date_all[1:]
    print(File_Date_all)
    num_Files = len(File_Date_all)

    # East Team Ranking Figure
    f = plt.figure()
    ax = f.add_subplot(111)
    plt.plot(Rank_all_TRR[1:], 'ro-', label='Toronto Raptors')
    plt.plot(Rank_all_MWB[1:], 'go-', label='Milwaukee Bucks')
    plt.plot(Rank_all_IDP[1:], 'bo-', label='Indiana Pacers')
    plt.plot(Rank_all_P76[1:], 'r*-', label='Philadelphia 76ers')
    plt.plot(Rank_all_DPT[1:], 'g*-', label='Detroit Pistons')
    plt.plot(Rank_all_OLM[1:], 'b*-', label='Orlando Magics')
    plt.plot(Rank_all_BOC[1:], 'rx-', label='Boston Celtics')
    plt.plot(Rank_all_CLH[1:], 'gx-', label='Charlotte Hornets')
    plt.plot(Rank_all_BRN[1:], 'bx-', label='Brooklyn Nets')
    plt.plot(Rank_all_MIH[1:], 'rd-', label='Miami Heats')
    plt.plot(Rank_all_WSW[1:], 'gd-', label='Washington Wizards')
    plt.plot(Rank_all_CCB[1:], 'bd-', label='Chicago Bulls')
    plt.plot(Rank_all_NYN[1:], 'rs-', label='New York Knicks')
    plt.plot(Rank_all_ATH[1:], 'gs-', label='Atlanta Hawks')
    plt.plot(Rank_all_CLC[1:], 'bs-', label='Cleveland Cavaliers')
    plt.title('East Teams Ranking History')
    plt.xlabel('Time')
    plt.ylabel('Rank #')
    ax.yaxis.set_ticks_position('both')
    leg = plt.legend(loc='upper left', fancybox=True, fontsize=9)
    leg.get_frame().set_alpha(0.2)
    # Draw a separate line for upper 8 teams and lower 8 teams
    x = np.arange(0, num_Files, 1)
    y = np.repeat(8.5, num_Files)
    plt.plot(x, y, 'k--')
    # for x_tick in range(0, num_Files):
    #     plt.text(x_tick, 1, File_Date_all[x_tick], rotation=45)

    # West Team Ranking Figure
    f = plt.figure()
    ax = f.add_subplot(111)
    plt.plot(Rank_all_PTB[1:], 'co-', label='Portland Trail Blazers')
    plt.plot(Rank_all_MGR[1:], 'yo-', label='Memphis Grizzlies')
    plt.plot(Rank_all_LAC[1:], 'ko-', label='LA Clippers')
    plt.plot(Rank_all_GSW[1:], 'c*-', label='Golden State Warriors')
    plt.plot(Rank_all_OKC[1:], 'y*-', label='Oklahoma City Thunders')
    plt.plot(Rank_all_DVN[1:], 'k*-', label='Denver Nuggets')
    plt.plot(Rank_all_NOP[1:], 'cx-', label='New Orleans Pelicans')
    plt.plot(Rank_all_LAL[1:], 'yx-', label='Los Angeles Lakers')
    plt.plot(Rank_all_HOR[1:], 'kx-', label='Houston Rockets')
    plt.plot(Rank_all_SAK[1:], 'cd-', label='Sacramento Kings')
    plt.plot(Rank_all_SAS[1:], 'yd-', label='San Antonio Spurs')
    plt.plot(Rank_all_UTJ[1:], 'kd-', label='Utah Jazz')
    plt.plot(Rank_all_DAM[1:], 'cs-', label='Dallas Mavericks')
    plt.plot(Rank_all_MTW[1:], 'ys-', label='Minnesota Timberwolves')
    plt.plot(Rank_all_PHS[1:], 'ks-', label='Phoenix Suns')
    plt.title('West Teams Ranking History')
    plt.xlabel('Time')
    plt.ylabel('Rank #')
    ax.yaxis.set_ticks_position('both')
    leg = plt.legend(loc='upper left', fancybox=True, fontsize=9)
    leg.get_frame().set_alpha(0.2)
    # Draw a separate line for upper 8 teams and lower 8 teams
    x = np.arange(0, num_Files, 1)
    y = np.repeat(8.5, num_Files)
    plt.plot(x, y, 'k--')

    # for x_tick in range(0, num_Files):
    #     plt.text(x_tick, 1, File_Date_all[x_tick], rotation=45)


    return
