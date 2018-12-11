import tkinter as tk
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import os, sys
from lxml import html
import requests
import xlsxwriter
import datetime


def function_Update():

    # Crawl from the website
    page = requests.get('https://nba.hupu.com/standings')
    tree = html.fromstring(page.content)

    # This will create a list of teams:
    teams = tree.xpath('//td/text()')
    names = tree.xpath('//a[@target="_blank"]/text()')

    # ==============================================
    # Define parameters
    teams_header_east = teams[3:15]
    teams_header_west = teams[205:217]
    NBA_East_Name = names[55:70]
    NBA_West_Name = names[70:85]
    rank = np.arange(0, 16, 1)
    rank = np.transpose(np.matrix(rank))
    m = len(teams)

    NBA_East = []
    NBA_West = []
    a = 0

    # -------------------------------------------
    # Store results into variables
    NBA_East = np.append(NBA_East, teams_header_east)
    NBA_West = np.append(NBA_West, teams_header_west)

    for idx in range(15, 111):
        if (((idx % 12) - 3) == 0):
            print('Teams: ', teams[idx:(idx + 12)])
            NBA_East = np.append(NBA_East, teams[idx:(idx + 12)])

    for idx in range(112, 203):
        if (((idx % 13) - 8) == 0):
            print('Teams: ', teams[idx:(idx + 12)])
            NBA_East = np.append(NBA_East, teams[idx:(idx + 12)])

    for idx in range(217, 313):
        if (((idx % 12) - 1) == 0):
            print('Teams: ', teams[idx:(idx + 12)])
            NBA_West = np.append(NBA_West, teams[idx:(idx + 12)])

    for idx in range(314, m):
        if (((idx % 13) - 2) == 0):
            print('Teams: ', teams[idx:(idx + 12)])
            NBA_West = np.append(NBA_West, teams[idx:(idx + 12)])

    NBA_East = np.matrix(NBA_East)
    NBA_East = np.reshape(NBA_East, (16, 12))
    NBA_West = np.matrix(NBA_West)
    NBA_West = np.reshape(NBA_West, (16, 12))

    # Get NBA team names into the result matrix
    NBA_East_Name = np.matrix(NBA_East_Name)
    NBA_East_Name = np.insert(NBA_East_Name, 0, a)
    NBA_East_Name = np.transpose(NBA_East_Name)
    NBA_West_Name = np.matrix(NBA_West_Name)
    NBA_West_Name = np.insert(NBA_West_Name, 0, a)
    NBA_West_Name = np.transpose(NBA_West_Name)
    East = np.hstack((NBA_East_Name, NBA_East))

    East = np.hstack((rank, East))
    West = np.hstack((NBA_West_Name, NBA_West))
    West = np.hstack((rank, West))

    # ============================================
    # Save the matrix variable results in Excel file
    now = datetime.datetime.now().date()
    filename = './NBA_Data/' + str(now) + '.xlsx'

    workbook = xlsxwriter.Workbook(filename)
    worksheet1 = workbook.add_worksheet()
    worksheet2 = workbook.add_worksheet()

    (p, q) = np.shape(East)

    for i in range(0, p):
        for j in range(0, q):
            worksheet1.write(i, j, East[i, j])
            worksheet2.write(i, j, West[i, j])

    workbook.close()

    return


def function_Plot():

    # ========================================
    # Define parameters
    Rank = 1
    Team = 2
    Wins = 3
    Loss = 4
    Win_Pctg = 5
    Win_Diff = 6
    Home_Result = 7
    Road_Result = 8
    Div_Result = 9
    Conf_Result = 10
    Pt_Gain = 11
    Pt_Loss = 12
    Pt_Diff = 13
    WL_Streak = 14

    if (val_Rank.get() == 1):
        selected_Column = Rank
    elif (val_Pt_Gain.get() == 1):
        selected_Column = Pt_Gain
    elif (val_Pt_Loss.get() == 1):
        selected_Column = Pt_Loss
    elif (val_Pt_Diff.get() == 1):
        selected_Column = Pt_Diff
    elif (val_WinPctg.get() == 1):
        selected_Column = Win_Pctg
    elif (val_HomeWinPctg.get() == 1):
        selected_Column = Home_Result
    elif (val_RoadWinPctg.get() == 1):
        selected_Column = Road_Result


    # Define parameters
    # East Teams
    Rank_all_TRR = 0
    Rank_all_MWB = 0
    Rank_all_IDP = 0
    Rank_all_P76 = 0
    Rank_all_DPT = 0
    Rank_all_OLM = 0
    Rank_all_BOC = 0
    Rank_all_CLH = 0
    Rank_all_BRN = 0
    Rank_all_MIH = 0
    Rank_all_WSW = 0
    Rank_all_CCB = 0
    Rank_all_NYN = 0
    Rank_all_ATH = 0
    Rank_all_CLC = 0

    # West Teams
    Rank_all_PTB = 0
    Rank_all_MGR = 0
    Rank_all_LAC = 0
    Rank_all_GSW = 0
    Rank_all_OKC = 0
    Rank_all_DVN = 0
    Rank_all_NOP = 0
    Rank_all_LAL = 0
    Rank_all_HOR = 0
    Rank_all_SAK = 0
    Rank_all_SAS = 0
    Rank_all_UTJ = 0
    Rank_all_DAM = 0
    Rank_all_MTW = 0
    Rank_all_PHS = 0

    # ========================================
    # Open a file
    path = "./NBA_Data/"
    dirs = os.listdir(path)

    # ========================================
    # Sweep all files in the folder
    for filename in dirs:
        filepath = './NBA_Data/' + filename
        wb = openpyxl.load_workbook(filepath)
        sheet1 = wb['Sheet1']
        sheet2 = wb['Sheet2']

        if (selected_Column == Rank) or \
                (selected_Column == Pt_Gain) or \
                (selected_Column == Pt_Loss) or \
                (selected_Column == Pt_Diff):

            # East Conference Ranking
            for i in range(1, 16):
                if ((sheet1.cell(row=i + 1, column=Team).value) == '猛龙') and (val_Raptors.get() == 1):
                    Rank_val_TRR = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_TRR = float(Rank_val_TRR)
                    Rank_all_TRR = np.append(Rank_all_TRR, Rank_val_TRR)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '雄鹿') and (val_Bucks.get() == 1):
                    Rank_val_MWB = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_MWB = float(Rank_val_MWB)
                    Rank_all_MWB = np.append(Rank_all_MWB, Rank_val_MWB)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '步行者') and (val_Pacers.get() == 1):
                    Rank_val_IDP = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_IDP = float(Rank_val_IDP)
                    Rank_all_IDP = np.append(Rank_all_IDP, Rank_val_IDP)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '76人') and (val_76ers.get() == 1):
                    Rank_val_P76 = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_P76 = float(Rank_val_P76)
                    Rank_all_P76 = np.append(Rank_all_P76, Rank_val_P76)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '活塞') and (val_Pistons.get() == 1):
                    Rank_val_DPT = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_DPT = float(Rank_val_DPT)
                    Rank_all_DPT = np.append(Rank_all_DPT, Rank_val_DPT)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '魔术') and (val_Magics.get() == 1):
                    Rank_val_OLM = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_OLM = float(Rank_val_OLM)
                    Rank_all_OLM = np.append(Rank_all_OLM, Rank_val_OLM)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '凯尔特人') and (val_Celtics.get() == 1):
                    Rank_val_BOC = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_BOC = float(Rank_val_BOC)
                    Rank_all_BOC = np.append(Rank_all_BOC, Rank_val_BOC)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '黄蜂') and (val_Hornets.get() == 1):
                    Rank_val_CLH = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_CLH = float(Rank_val_CLH)
                    Rank_all_CLH = np.append(Rank_all_CLH, Rank_val_CLH)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '篮网') and (val_Nets.get() == 1):
                    Rank_val_BRN = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_BRN = float(Rank_val_BRN)
                    Rank_all_BRN = np.append(Rank_all_BRN, Rank_val_BRN)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '热火') and (val_Heats.get() == 1):
                    Rank_val_MIH = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_MIH = float(Rank_val_MIH)
                    Rank_all_MIH = np.append(Rank_all_MIH, Rank_val_MIH)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '奇才') and (val_Wizards.get() == 1):
                    Rank_val_WSW = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_WSW = float(Rank_val_WSW)
                    Rank_all_WSW = np.append(Rank_all_WSW, Rank_val_WSW)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '公牛') and (val_Bulls.get() == 1):
                    Rank_val_CCB = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_CCB = float(Rank_val_CCB)
                    Rank_all_CCB = np.append(Rank_all_CCB, Rank_val_CCB)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '尼克斯') and (val_Knicks.get() == 1):
                    Rank_val_NYN = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_NYN = float(Rank_val_NYN)
                    Rank_all_NYN = np.append(Rank_all_NYN, Rank_val_NYN)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '老鹰') and (val_Hawks.get() == 1):
                    Rank_val_ATH = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_ATH = float(Rank_val_ATH)
                    Rank_all_ATH = np.append(Rank_all_ATH, Rank_val_ATH)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '骑士') and (val_Cavaliers.get() == 1):
                    Rank_val_CLC = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_CLC = float(Rank_val_CLC)
                    Rank_all_CLC = np.append(Rank_all_CLC, Rank_val_CLC)

            # West Conference Ranking
            for i in range(1, 16):
                if ((sheet2.cell(row=i + 1, column=Team).value) == '开拓者') and (val_Blazers.get() == 1):
                    Rank_val_PTB = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_PTB = float(Rank_val_PTB)
                    Rank_all_PTB = np.append(Rank_all_PTB, Rank_val_PTB)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '灰熊') and (val_Grizzlies.get() == 1):
                    Rank_val_MGR = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_MGR = float(Rank_val_MGR)
                    Rank_all_MGR = np.append(Rank_all_MGR, Rank_val_MGR)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '快船') and (val_Clippers.get() == 1):
                    Rank_val_LAC = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_LAC = float(Rank_val_LAC)
                    Rank_all_LAC = np.append(Rank_all_LAC, Rank_val_LAC)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '勇士') and (val_Warriors.get() == 1):
                    Rank_val_GSW = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_GSW = float(Rank_val_GSW)
                    Rank_all_GSW = np.append(Rank_all_GSW, Rank_val_GSW)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '雷霆') and (val_Thunders.get() == 1):
                    Rank_val_OKC = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_OKC = float(Rank_val_OKC)
                    Rank_all_OKC = np.append(Rank_all_OKC, Rank_val_OKC)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '掘金') and (val_Nuggets.get() == 1):
                    Rank_val_DVN = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_DVN = float(Rank_val_DVN)
                    Rank_all_DVN = np.append(Rank_all_DVN, Rank_val_DVN)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '鹈鹕') and (val_Pelicans.get() == 1):
                    Rank_val_NOP = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_NOP = float(Rank_val_NOP)
                    Rank_all_NOP = np.append(Rank_all_NOP, Rank_val_NOP)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '湖人') and (val_Lakers.get() == 1):
                    Rank_val_LAL = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_LAL = float(Rank_val_LAL)
                    Rank_all_LAL = np.append(Rank_all_LAL, Rank_val_LAL)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '火箭') and (val_Rockets.get() == 1):
                    Rank_val_HOR = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_HOR = float(Rank_val_HOR)
                    Rank_all_HOR = np.append(Rank_all_HOR, Rank_val_HOR)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '国王') and (val_Kings.get() == 1):
                    Rank_val_SAK = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_SAK = float(Rank_val_SAK)
                    Rank_all_SAK = np.append(Rank_all_SAK, Rank_val_SAK)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '马刺') and (val_Spurs.get() == 1):
                    Rank_val_SAS = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_SAS = float(Rank_val_SAS)
                    Rank_all_SAS = np.append(Rank_all_SAS, Rank_val_SAS)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '爵士') and (val_Jazz.get() == 1):
                    Rank_val_UTJ = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_UTJ = float(Rank_val_UTJ)
                    Rank_all_UTJ = np.append(Rank_all_UTJ, Rank_val_UTJ)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '独行侠') and (val_Mavericks.get() == 1):
                    Rank_val_DAM = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_DAM = float(Rank_val_DAM)
                    Rank_all_DAM = np.append(Rank_all_DAM, Rank_val_DAM)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '森林狼') and (val_Timberwolves.get() == 1):
                    Rank_val_MTW = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_MTW = float(Rank_val_MTW)
                    Rank_all_MTW = np.append(Rank_all_MTW, Rank_val_MTW)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '太阳') and (val_Suns.get() == 1):
                    Rank_val_PHS = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_PHS = float(Rank_val_PHS)
                    Rank_all_PHS = np.append(Rank_all_PHS, Rank_val_PHS)


        elif (selected_Column == Home_Result) or \
                (selected_Column == Road_Result):

            # East Conference Ranking
            for i in range(1, 16):
                if ((sheet1.cell(row=i + 1, column=Team).value) == '猛龙') and (val_Raptors.get() == 1):
                    Rank_val_TRR = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_TRR = Rank_val_TRR.split('-')
                    Rank_val_TRR_W = float(Rank_val_TRR[0])
                    Rank_val_TRR_L = float(Rank_val_TRR[1])
                    if ((Rank_val_TRR_W == 0) and (Rank_val_TRR_L == 0)):
                        Rank_val_TRR = 0
                    else:
                        Rank_val_TRR = 100 * Rank_val_TRR_W / (Rank_val_TRR_W + Rank_val_TRR_L)
                    Rank_all_TRR = np.append(Rank_all_TRR, Rank_val_TRR)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '雄鹿') and (val_Bucks.get() == 1):
                    Rank_val_MWB = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_MWB = Rank_val_MWB.split('-')
                    Rank_val_MWB_W = float(Rank_val_MWB[0])
                    Rank_val_MWB_L = float(Rank_val_MWB[1])
                    if ((Rank_val_MWB_W == 0) and (Rank_val_MWB_L == 0)):
                        Rank_val_MWB = 0
                    else:
                        Rank_val_MWB = 100 * Rank_val_MWB_W / (Rank_val_MWB_W + Rank_val_MWB_L)
                    Rank_all_MWB = np.append(Rank_all_MWB, Rank_val_MWB)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '步行者') and (val_Pacers.get() == 1):
                    Rank_val_IDP = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_IDP = Rank_val_IDP.split('-')
                    Rank_val_IDP_W = float(Rank_val_IDP[0])
                    Rank_val_IDP_L = float(Rank_val_IDP[1])
                    if ((Rank_val_IDP_W == 0) and (Rank_val_IDP_L == 0)):
                        Rank_val_IDP = 0
                    else:
                        Rank_val_IDP = 100 * Rank_val_IDP_W / (Rank_val_IDP_W + Rank_val_IDP_L)
                    Rank_all_IDP = np.append(Rank_all_IDP, Rank_val_IDP)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '76人') and (val_76ers.get() == 1):
                    Rank_val_P76 = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_P76 = Rank_val_P76.split('-')
                    Rank_val_P76_W = float(Rank_val_P76[0])
                    Rank_val_P76_L = float(Rank_val_P76[1])
                    if ((Rank_val_P76_W == 0) and (Rank_val_P76_L == 0)):
                        Rank_val_P76 = 0
                    else:
                        Rank_val_P76 = 100 * Rank_val_P76_W / (Rank_val_P76_W + Rank_val_P76_L)
                    Rank_all_P76 = np.append(Rank_all_P76, Rank_val_P76)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '活塞') and (val_Pistons.get() == 1):
                    Rank_val_DPT = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_DPT = Rank_val_DPT.split('-')
                    Rank_val_DPT_W = float(Rank_val_DPT[0])
                    Rank_val_DPT_L = float(Rank_val_DPT[1])
                    if ((Rank_val_DPT_W == 0) and (Rank_val_DPT_L == 0)):
                        Rank_val_DPT = 0
                    else:
                        Rank_val_DPT = 100 * Rank_val_DPT_W / (Rank_val_DPT_W + Rank_val_DPT_L)
                    Rank_all_DPT = np.append(Rank_all_DPT, Rank_val_DPT)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '魔术') and (val_Magics.get() == 1):
                    Rank_val_OLM = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_OLM = Rank_val_OLM.split('-')
                    Rank_val_OLM_W = float(Rank_val_OLM[0])
                    Rank_val_OLM_L = float(Rank_val_OLM[1])
                    if ((Rank_val_OLM_W == 0) and (Rank_val_OLM_L == 0)):
                        Rank_val_OLM = 0
                    else:
                        Rank_val_OLM = 100 * Rank_val_OLM_W / (Rank_val_OLM_W + Rank_val_OLM_L)
                    Rank_all_OLM = np.append(Rank_all_OLM, Rank_val_OLM)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '凯尔特人') and (val_Celtics.get() == 1):
                    Rank_val_BOC = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_BOC = Rank_val_BOC.split('-')
                    Rank_val_BOC_W = float(Rank_val_BOC[0])
                    Rank_val_BOC_L = float(Rank_val_BOC[1])
                    if ((Rank_val_BOC_W == 0) and (Rank_val_BOC_L == 0)):
                        Rank_val_BOC = 0
                    else:
                        Rank_val_BOC = 100 * Rank_val_BOC_W / (Rank_val_BOC_W + Rank_val_BOC_L)
                    Rank_all_BOC = np.append(Rank_all_BOC, Rank_val_BOC)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '黄蜂') and (val_Hornets.get() == 1):
                    Rank_val_CLH = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_CLH = Rank_val_CLH.split('-')
                    Rank_val_CLH_W = float(Rank_val_CLH[0])
                    Rank_val_CLH_L = float(Rank_val_CLH[1])
                    if ((Rank_val_CLH_W == 0) and (Rank_val_CLH_L == 0)):
                        Rank_val_CLH = 0
                    else:
                        Rank_val_CLH = 100 * Rank_val_CLH_W / (Rank_val_CLH_W + Rank_val_CLH_L)
                    Rank_all_CLH = np.append(Rank_all_CLH, Rank_val_CLH)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '篮网') and (val_Nets.get() == 1):
                    Rank_val_BRN = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_BRN = Rank_val_BRN.split('-')
                    Rank_val_BRN_W = float(Rank_val_BRN[0])
                    Rank_val_BRN_L = float(Rank_val_BRN[1])
                    if ((Rank_val_BRN_W == 0) and (Rank_val_BRN_L == 0)):
                        Rank_val_BRN = 0
                    else:
                        Rank_val_BRN = 100 * Rank_val_BRN_W / (Rank_val_BRN_W + Rank_val_BRN_L)
                    Rank_all_BRN = np.append(Rank_all_BRN, Rank_val_BRN)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '热火') and (val_Heats.get() == 1):
                    Rank_val_MIH = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_MIH = Rank_val_MIH.split('-')
                    Rank_val_MIH_W = float(Rank_val_MIH[0])
                    Rank_val_MIH_L = float(Rank_val_MIH[1])
                    if ((Rank_val_MIH_W == 0) and (Rank_val_MIH_L == 0)):
                        Rank_val_MIH = 0
                    else:
                        Rank_val_MIH = 100 * Rank_val_MIH_W / (Rank_val_MIH_W + Rank_val_MIH_L)
                    Rank_all_MIH = np.append(Rank_all_MIH, Rank_val_MIH)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '奇才') and (val_Wizards.get() == 1):
                    Rank_val_WSW = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_WSW = Rank_val_WSW.split('-')
                    Rank_val_WSW_W = float(Rank_val_WSW[0])
                    Rank_val_WSW_L = float(Rank_val_WSW[1])
                    if ((Rank_val_WSW_W == 0) and (Rank_val_WSW_L == 0)):
                        Rank_val_WSW = 0
                    else:
                        Rank_val_WSW = 100 * Rank_val_WSW_W / (Rank_val_WSW_W + Rank_val_WSW_L)
                    Rank_all_WSW = np.append(Rank_all_WSW, Rank_val_WSW)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '公牛') and (val_Bulls.get() == 1):
                    Rank_val_CCB = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_CCB = Rank_val_CCB.split('-')
                    Rank_val_CCB_W = float(Rank_val_CCB[0])
                    Rank_val_CCB_L = float(Rank_val_CCB[1])
                    if ((Rank_val_CCB_W == 0) and (Rank_val_CCB_L == 0)):
                        Rank_val_CCB = 0
                    else:
                        Rank_val_CCB = 100 * Rank_val_CCB_W / (Rank_val_CCB_W + Rank_val_CCB_L)
                    Rank_all_CCB = np.append(Rank_all_CCB, Rank_val_CCB)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '尼克斯') and (val_Knicks.get() == 1):
                    Rank_val_NYN = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_NYN = Rank_val_NYN.split('-')
                    Rank_val_NYN_W = float(Rank_val_NYN[0])
                    Rank_val_NYN_L = float(Rank_val_NYN[1])
                    if ((Rank_val_NYN_W == 0) and (Rank_val_NYN_L == 0)):
                        Rank_val_NYN = 0
                    else:
                        Rank_val_NYN = 100 * Rank_val_NYN_W / (Rank_val_NYN_W + Rank_val_NYN_L)
                    Rank_all_NYN = np.append(Rank_all_NYN, Rank_val_NYN)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '老鹰') and (val_Hawks.get() == 1):
                    Rank_val_ATH = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_ATH = Rank_val_ATH.split('-')
                    Rank_val_ATH_W = float(Rank_val_ATH[0])
                    Rank_val_ATH_L = float(Rank_val_ATH[1])
                    if ((Rank_val_ATH_W == 0) and (Rank_val_ATH_L == 0)):
                        Rank_val_ATH = 0
                    else:
                        Rank_val_ATH = 100 * Rank_val_ATH_W / (Rank_val_ATH_W + Rank_val_ATH_L)
                    Rank_all_ATH = np.append(Rank_all_ATH, Rank_val_ATH)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '骑士') and (val_Cavaliers.get() == 1):
                    Rank_val_CLC = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_CLC = Rank_val_CLC.split('-')
                    Rank_val_CLC_W = float(Rank_val_CLC[0])
                    Rank_val_CLC_L = float(Rank_val_CLC[1])
                    if ((Rank_val_CLC_W == 0) and (Rank_val_CLC_L == 0)):
                        Rank_val_CLC = 0
                    else:
                        Rank_val_CLC = 100 * Rank_val_CLC_W / (Rank_val_CLC_W + Rank_val_CLC_L)
                    Rank_all_CLC = np.append(Rank_all_CLC, Rank_val_CLC)

            # West Conference Ranking
            for i in range(1, 16):
                if ((sheet2.cell(row=i + 1, column=Team).value) == '开拓者') and (val_Blazers.get() == 1):
                    Rank_val_PTB = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_PTB = Rank_val_PTB.split('-')
                    Rank_val_PTB_W = float(Rank_val_PTB[0])
                    Rank_val_PTB_L = float(Rank_val_PTB[1])
                    if ((Rank_val_PTB_W == 0) and (Rank_val_PTB_L == 0)):
                        Rank_val_PTB = 0
                    else:
                        Rank_val_PTB = 100 * Rank_val_PTB_W / (Rank_val_PTB_W + Rank_val_PTB_L)
                    Rank_all_PTB = np.append(Rank_all_PTB, Rank_val_PTB)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '灰熊') and (val_Grizzlies.get() == 1):
                    Rank_val_MGR = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_MGR = Rank_val_MGR.split('-')
                    Rank_val_MGR_W = float(Rank_val_MGR[0])
                    Rank_val_MGR_L = float(Rank_val_MGR[1])
                    if ((Rank_val_MGR_W == 0) and (Rank_val_MGR_L == 0)):
                        Rank_val_MGR = 0
                    else:
                        Rank_val_MGR = 100 * Rank_val_MGR_W / (Rank_val_MGR_W + Rank_val_MGR_L)
                    Rank_all_MGR = np.append(Rank_all_MGR, Rank_val_MGR)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '快船') and (val_Clippers.get() == 1):
                    Rank_val_LAC = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_LAC = Rank_val_LAC.split('-')
                    Rank_val_LAC_W = float(Rank_val_LAC[0])
                    Rank_val_LAC_L = float(Rank_val_LAC[1])
                    if ((Rank_val_LAC_W == 0) and (Rank_val_LAC_L == 0)):
                        Rank_val_LAC = 0
                    else:
                        Rank_val_LAC = 100 * Rank_val_LAC_W / (Rank_val_LAC_W + Rank_val_LAC_L)
                    Rank_all_LAC = np.append(Rank_all_LAC, Rank_val_LAC)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '勇士') and (val_Warriors.get() == 1):
                    Rank_val_GSW = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_GSW = Rank_val_GSW.split('-')
                    Rank_val_GSW_W = float(Rank_val_GSW[0])
                    Rank_val_GSW_L = float(Rank_val_GSW[1])
                    if ((Rank_val_GSW_W == 0) and (Rank_val_GSW_L == 0)):
                        Rank_val_GSW = 0
                    else:
                        Rank_val_GSW = 100 * Rank_val_GSW_W / (Rank_val_GSW_W + Rank_val_GSW_L)
                    Rank_all_GSW = np.append(Rank_all_GSW, Rank_val_GSW)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '雷霆') and (val_Thunders.get() == 1):
                    Rank_val_OKC = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_OKC = Rank_val_OKC.split('-')
                    Rank_val_OKC_W = float(Rank_val_OKC[0])
                    Rank_val_OKC_L = float(Rank_val_OKC[1])
                    if ((Rank_val_OKC_W == 0) and (Rank_val_OKC_L == 0)):
                        Rank_val_OKC = 0
                    else:
                        Rank_val_OKC = 100 * Rank_val_OKC_W / (Rank_val_OKC_W + Rank_val_OKC_L)
                    Rank_all_OKC = np.append(Rank_all_OKC, Rank_val_OKC)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '掘金') and (val_Nuggets.get() == 1):
                    Rank_val_DVN = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_DVN = Rank_val_DVN.split('-')
                    Rank_val_DVN_W = float(Rank_val_DVN[0])
                    Rank_val_DVN_L = float(Rank_val_DVN[1])
                    if ((Rank_val_DVN_W == 0) and (Rank_val_DVN_L == 0)):
                        Rank_val_DVN = 0
                    else:
                        Rank_val_DVN = 100 * Rank_val_DVN_W / (Rank_val_DVN_W + Rank_val_DVN_L)
                    Rank_all_DVN = np.append(Rank_all_DVN, Rank_val_DVN)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '鹈鹕') and (val_Pelicans.get() == 1):
                    Rank_val_NOP = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_NOP = Rank_val_NOP.split('-')
                    Rank_val_NOP_W = float(Rank_val_NOP[0])
                    Rank_val_NOP_L = float(Rank_val_NOP[1])
                    if ((Rank_val_NOP_W == 0) and (Rank_val_NOP_L == 0)):
                        Rank_val_NOP = 0
                    else:
                        Rank_val_NOP = 100 * Rank_val_NOP_W / (Rank_val_NOP_W + Rank_val_NOP_L)
                    Rank_all_NOP = np.append(Rank_all_NOP, Rank_val_NOP)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '湖人') and (val_Lakers.get() == 1):
                    Rank_val_LAL = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_LAL = Rank_val_LAL.split('-')
                    Rank_val_LAL_W = float(Rank_val_LAL[0])
                    Rank_val_LAL_L = float(Rank_val_LAL[1])
                    if ((Rank_val_LAL_W == 0) and (Rank_val_LAL_L == 0)):
                        Rank_val_LAL = 0
                    else:
                        Rank_val_LAL = 100 * Rank_val_LAL_W / (Rank_val_LAL_W + Rank_val_LAL_L)
                    Rank_all_LAL = np.append(Rank_all_LAL, Rank_val_LAL)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '火箭') and (val_Rockets.get() == 1):
                    Rank_val_HOR = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_HOR = Rank_val_HOR.split('-')
                    Rank_val_HOR_W = float(Rank_val_HOR[0])
                    Rank_val_HOR_L = float(Rank_val_HOR[1])
                    if ((Rank_val_HOR_W == 0) and (Rank_val_HOR_L == 0)):
                        Rank_val_HOR = 0
                    else:
                        Rank_val_HOR = 100 * Rank_val_HOR_W / (Rank_val_HOR_W + Rank_val_HOR_L)
                    Rank_all_HOR = np.append(Rank_all_HOR, Rank_val_HOR)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '国王') and (val_Kings.get() == 1):
                    Rank_val_SAK = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_SAK = Rank_val_SAK.split('-')
                    Rank_val_SAK_W = float(Rank_val_SAK[0])
                    Rank_val_SAK_L = float(Rank_val_SAK[1])
                    if ((Rank_val_SAK_W == 0) and (Rank_val_SAK_L == 0)):
                        Rank_val_SAK = 0
                    else:
                        Rank_val_SAK = 100 * Rank_val_SAK_W / (Rank_val_SAK_W + Rank_val_SAK_L)
                    Rank_all_SAK = np.append(Rank_all_SAK, Rank_val_SAK)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '马刺') and (val_Spurs.get() == 1):
                    Rank_val_SAS = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_SAS = Rank_val_SAS.split('-')
                    Rank_val_SAS_W = float(Rank_val_SAS[0])
                    Rank_val_SAS_L = float(Rank_val_SAS[1])
                    if ((Rank_val_SAS_W == 0) and (Rank_val_SAS_L == 0)):
                        Rank_val_SAS = 0
                    else:
                        Rank_val_SAS = 100 * Rank_val_SAS_W / (Rank_val_SAS_W + Rank_val_SAS_L)
                    Rank_all_SAS = np.append(Rank_all_SAS, Rank_val_SAS)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '爵士') and (val_Jazz.get() == 1):
                    Rank_val_UTJ = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_UTJ = Rank_val_UTJ.split('-')
                    Rank_val_UTJ_W = float(Rank_val_UTJ[0])
                    Rank_val_UTJ_L = float(Rank_val_UTJ[1])
                    if ((Rank_val_UTJ_W == 0) and (Rank_val_UTJ_L == 0)):
                        Rank_val_UTJ = 0
                    else:
                        Rank_val_UTJ = 100 * Rank_val_UTJ_W / (Rank_val_UTJ_W + Rank_val_UTJ_L)
                    Rank_all_UTJ = np.append(Rank_all_UTJ, Rank_val_UTJ)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '独行侠') and (val_Mavericks.get() == 1):
                    Rank_val_DAM = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_DAM = Rank_val_DAM.split('-')
                    Rank_val_DAM_W = float(Rank_val_DAM[0])
                    Rank_val_DAM_L = float(Rank_val_DAM[1])
                    if ((Rank_val_DAM_W == 0) and (Rank_val_DAM_L == 0)):
                        Rank_val_DAM = 0
                    else:
                        Rank_val_DAM = 100 * Rank_val_DAM_W / (Rank_val_DAM_W + Rank_val_DAM_L)
                    Rank_all_DAM = np.append(Rank_all_DAM, Rank_val_DAM)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '森林狼') and (val_Timberwolves.get() == 1):
                    Rank_val_MTW = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_MTW = Rank_val_MTW.split('-')
                    Rank_val_MTW_W = float(Rank_val_MTW[0])
                    Rank_val_MTW_L = float(Rank_val_MTW[1])
                    if ((Rank_val_MTW_W == 0) and (Rank_val_MTW_L == 0)):
                        Rank_val_MTW = 0
                    else:
                        Rank_val_MTW = 100 * Rank_val_MTW_W / (Rank_val_MTW_W + Rank_val_MTW_L)
                    Rank_all_MTW = np.append(Rank_all_MTW, Rank_val_MTW)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '太阳') and (val_Suns.get() == 1):
                    Rank_val_PHS = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_PHS = Rank_val_PHS.split('-')
                    Rank_val_PHS_W = float(Rank_val_PHS[0])
                    Rank_val_PHS_L = float(Rank_val_PHS[1])
                    if ((Rank_val_PHS_W == 0) and (Rank_val_PHS_L == 0)):
                        Rank_val_PHS = 0
                    else:
                        Rank_val_PHS = 100 * Rank_val_PHS_W / (Rank_val_PHS_W + Rank_val_PHS_L)
                    Rank_all_PHS = np.append(Rank_all_PHS, Rank_val_PHS)

        elif (selected_Column == Win_Pctg):

            # East Conference Ranking
            for i in range(1, 16):
                if ((sheet1.cell(row=i + 1, column=Team).value) == '猛龙') and (val_Raptors.get() == 1):
                    Rank_val_TRR = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_TRR = float(Rank_val_TRR.strip('%'))
                    Rank_all_TRR = np.append(Rank_all_TRR, Rank_val_TRR)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '雄鹿') and (val_Bucks.get() == 1):
                    Rank_val_MWB = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_MWB = float(Rank_val_MWB.strip('%'))
                    Rank_all_MWB = np.append(Rank_all_MWB, Rank_val_MWB)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '步行者') and (val_Pacers.get() == 1):
                    Rank_val_IDP = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_IDP = float(Rank_val_IDP.strip('%'))
                    Rank_all_IDP = np.append(Rank_all_IDP, Rank_val_IDP)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '76人') and (val_76ers.get() == 1):
                    Rank_val_P76 = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_P76 = float(Rank_val_P76.strip('%'))
                    Rank_all_P76 = np.append(Rank_all_P76, Rank_val_P76)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '活塞') and (val_Pistons.get() == 1):
                    Rank_val_DPT = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_DPT = float(Rank_val_DPT.strip('%'))
                    Rank_all_DPT = np.append(Rank_all_DPT, Rank_val_DPT)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '魔术') and (val_Magics.get() == 1):
                    Rank_val_OLM = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_OLM = float(Rank_val_OLM.strip('%'))
                    Rank_all_OLM = np.append(Rank_all_OLM, Rank_val_OLM)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '凯尔特人') and (val_Celtics.get() == 1):
                    Rank_val_BOC = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_BOC = float(Rank_val_BOC.strip('%'))
                    Rank_all_BOC = np.append(Rank_all_BOC, Rank_val_BOC)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '黄蜂') and (val_Hornets.get() == 1):
                    Rank_val_CLH = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_CLH = float(Rank_val_CLH.strip('%'))
                    Rank_all_CLH = np.append(Rank_all_CLH, Rank_val_CLH)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '篮网') and (val_Nets.get() == 1):
                    Rank_val_BRN = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_BRN = float(Rank_val_BRN.strip('%'))
                    Rank_all_BRN = np.append(Rank_all_BRN, Rank_val_BRN)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '热火') and (val_Heats.get() == 1):
                    Rank_val_MIH = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_MIH = float(Rank_val_MIH.strip('%'))
                    Rank_all_MIH = np.append(Rank_all_MIH, Rank_val_MIH)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '奇才') and (val_Wizards.get() == 1):
                    Rank_val_WSW = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_WSW = float(Rank_val_WSW.strip('%'))
                    Rank_all_WSW = np.append(Rank_all_WSW, Rank_val_WSW)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '公牛') and (val_Bulls.get() == 1):
                    Rank_val_CCB = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_CCB = float(Rank_val_CCB.strip('%'))
                    Rank_all_CCB = np.append(Rank_all_CCB, Rank_val_CCB)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '尼克斯') and (val_Knicks.get() == 1):
                    Rank_val_NYN = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_NYN = float(Rank_val_NYN.strip('%'))
                    Rank_all_NYN = np.append(Rank_all_NYN, Rank_val_NYN)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '老鹰') and (val_Hawks.get() == 1):
                    Rank_val_ATH = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_ATH = float(Rank_val_ATH.strip('%'))
                    Rank_all_ATH = np.append(Rank_all_ATH, Rank_val_ATH)

                elif ((sheet1.cell(row=i + 1, column=Team).value) == '骑士') and (val_Cavaliers.get() == 1):
                    Rank_val_CLC = sheet1.cell(row=i + 1, column=selected_Column).value
                    Rank_val_CLC = float(Rank_val_CLC.strip('%'))
                    Rank_all_CLC = np.append(Rank_all_CLC, Rank_val_CLC)

            # West Conference Ranking
            for i in range(1, 16):
                if ((sheet2.cell(row=i + 1, column=Team).value) == '开拓者') and (val_Blazers.get() == 1):
                    Rank_val_PTB = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_PTB = float(Rank_val_PTB.strip('%'))
                    Rank_all_PTB = np.append(Rank_all_PTB, Rank_val_PTB)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '灰熊') and (val_Grizzlies.get() == 1):
                    Rank_val_MGR = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_MGR = float(Rank_val_MGR.strip('%'))
                    Rank_all_MGR = np.append(Rank_all_MGR, Rank_val_MGR)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '快船') and (val_Clippers.get() == 1):
                    Rank_val_LAC = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_LAC = float(Rank_val_LAC.strip('%'))
                    Rank_all_LAC = np.append(Rank_all_LAC, Rank_val_LAC)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '勇士') and (val_Warriors.get() == 1):
                    Rank_val_GSW = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_GSW = float(Rank_val_GSW.strip('%'))
                    Rank_all_GSW = np.append(Rank_all_GSW, Rank_val_GSW)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '雷霆') and (val_Thunders.get() == 1):
                    Rank_val_OKC = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_OKC = float(Rank_val_OKC.strip('%'))
                    Rank_all_OKC = np.append(Rank_all_OKC, Rank_val_OKC)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '掘金') and (val_Nuggets.get() == 1):
                    Rank_val_DVN = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_DVN = float(Rank_val_DVN.strip('%'))
                    Rank_all_DVN = np.append(Rank_all_DVN, Rank_val_DVN)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '鹈鹕') and (val_Pelicans.get() == 1):
                    Rank_val_NOP = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_NOP = float(Rank_val_NOP.strip('%'))
                    Rank_all_NOP = np.append(Rank_all_NOP, Rank_val_NOP)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '湖人') and (val_Lakers.get() == 1):
                    Rank_val_LAL = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_LAL = float(Rank_val_LAL.strip('%'))
                    Rank_all_LAL = np.append(Rank_all_LAL, Rank_val_LAL)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '火箭') and (val_Rockets.get() == 1):
                    Rank_val_HOR = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_HOR = float(Rank_val_HOR.strip('%'))
                    Rank_all_HOR = np.append(Rank_all_HOR, Rank_val_HOR)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '国王') and (val_Kings.get() == 1):
                    Rank_val_SAK = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_SAK = float(Rank_val_SAK.strip('%'))
                    Rank_all_SAK = np.append(Rank_all_SAK, Rank_val_SAK)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '马刺') and (val_Spurs.get() == 1):
                    Rank_val_SAS = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_SAS = float(Rank_val_SAS.strip('%'))
                    Rank_all_SAS = np.append(Rank_all_SAS, Rank_val_SAS)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '爵士') and (val_Jazz.get() == 1):
                    Rank_val_UTJ = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_UTJ = float(Rank_val_UTJ.strip('%'))
                    Rank_all_UTJ = np.append(Rank_all_UTJ, Rank_val_UTJ)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '独行侠') and (val_Mavericks.get() == 1):
                    Rank_val_DAM = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_DAM = float(Rank_val_DAM.strip('%'))
                    Rank_all_DAM = np.append(Rank_all_DAM, Rank_val_DAM)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '森林狼') and (val_Timberwolves.get() == 1):
                    Rank_val_MTW = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_MTW = float(Rank_val_MTW.strip('%'))
                    Rank_all_MTW = np.append(Rank_all_MTW, Rank_val_MTW)

                elif ((sheet2.cell(row=i + 1, column=Team).value) == '太阳') and (val_Suns.get() == 1):
                    Rank_val_PHS = sheet2.cell(row=i + 1, column=selected_Column).value
                    Rank_val_PHS = float(Rank_val_PHS.strip('%'))
                    Rank_all_PHS = np.append(Rank_all_PHS, Rank_val_PHS)



    showMatrix = [val_Raptors.get(),
                  val_Bucks.get(),
                  val_Pacers.get(),
                  val_76ers.get(),
                  val_Pistons.get(),
                  val_Magics.get(),
                  val_Celtics.get(),
                  val_Hornets.get(),
                  val_Nets.get(),
                  val_Heats.get(),
                  val_Wizards.get(),
                  val_Bulls.get(),
                  val_Knicks.get(),
                  val_Hawks.get(),
                  val_Cavaliers.get(),
                  val_Blazers.get(),
                  val_Grizzlies.get(),
                  val_Clippers.get(),
                  val_Warriors.get(),
                  val_Thunders.get(),
                  val_Nuggets.get(),
                  val_Pelicans.get(),
                  val_Lakers.get(),
                  val_Rockets.get(),
                  val_Kings.get(),
                  val_Spurs.get(),
                  val_Jazz.get(),
                  val_Mavericks.get(),
                  val_Timberwolves.get(),
                  val_Suns.get()]

    f = plt.figure()
    for i in range(0, 30):
        if (showMatrix[i] == 1):
            if (i == 0):
                plt.plot(Rank_all_TRR[1:], 'ro-', label='Toronto Raptors')
            elif (i == 1):
                plt.plot(Rank_all_MWB[1:], 'go-', label='Milwaukee Bucks')
            elif (i == 2):
                plt.plot(Rank_all_IDP[1:], 'bo-', label='Indiana Pacers')
            elif (i == 3):
                plt.plot(Rank_all_P76[1:], 'r*-', label='Philadelphia 76ers')
            elif (i == 4):
                plt.plot(Rank_all_DPT[1:], 'g*-', label='Detroit Pistons')
            elif (i == 5):
                plt.plot(Rank_all_OLM[1:], 'b*-', label='Orlando Magics')
            elif (i == 6):
                plt.plot(Rank_all_BOC[1:], 'rx-', label='Boston Celtics')
            elif (i == 7):
                plt.plot(Rank_all_CLH[1:], 'gx-', label='Charlotte Hornets')
            elif (i == 8):
                plt.plot(Rank_all_BRN[1:], 'bx-', label='Brooklyn Nets')
            elif (i == 9):
                plt.plot(Rank_all_MIH[1:], 'rd-', label='Miami Heats')
            elif (i == 10):
                plt.plot(Rank_all_WSW[1:], 'gd-', label='Washington Wizards')
            elif (i == 11):
                plt.plot(Rank_all_CCB[1:], 'bd-', label='Chicago Bulls')
            elif (i == 12):
                plt.plot(Rank_all_NYN[1:], 'rs-', label='New York Knicks')
            elif (i == 13):
                plt.plot(Rank_all_ATH[1:], 'gs-', label='Atlanta Hawks')
            elif (i == 14):
                plt.plot(Rank_all_CLC[1:], 'bs-', label='Cleveland Cavaliers')
            elif (i == 15):
                plt.plot(Rank_all_PTB[1:], 'co-', label='Portland Trail Blazers')
            elif (i == 16):
                plt.plot(Rank_all_MGR[1:], 'yo-', label='Memphis Grizzlies')
            elif (i == 17):
                plt.plot(Rank_all_LAC[1:], 'ko-', label='LA Clippers')
            elif (i == 18):
                plt.plot(Rank_all_GSW[1:], 'c*-', label='Golden State Warriors')
            elif (i == 19):
                plt.plot(Rank_all_OKC[1:], 'y*-', label='Oklahoma City Thunders')
            elif (i == 20):
                plt.plot(Rank_all_DVN[1:], 'k*-', label='Denver Nuggets')
            elif (i == 21):
                plt.plot(Rank_all_NOP[1:], 'cx-', label='New Orleans Pelicans')
            elif (i == 22):
                plt.plot(Rank_all_LAL[1:], 'yx-', label='Los Angeles Lakers')
            elif (i == 23):
                plt.plot(Rank_all_HOR[1:], 'kx-', label='Houston Rockets')
            elif (i == 24):
                plt.plot(Rank_all_SAK[1:], 'cd-', label='Sacramento Kings')
            elif (i == 25):
                plt.plot(Rank_all_SAS[1:], 'yd-', label='San Antonio Spurs')
            elif (i == 26):
                plt.plot(Rank_all_UTJ[1:], 'kd-', label='Utah Jazz')
            elif (i == 27):
                plt.plot(Rank_all_DAM[1:], 'cs-', label='Dallas Mavericks')
            elif (i == 28):
                plt.plot(Rank_all_MTW[1:], 'ys-', label='Minnesota Timberwolves')
            elif (i == 29):
                plt.plot(Rank_all_PHS[1:], 'ks-', label='Phoenix Suns')

    plt.legend(loc='upper left', fancybox=True, framealpha=0.1)
    plt.show()

    return




# =================================
# Define parameters
master = tk.Tk()
master.title('NBA Hupu Data Visualization')
master.geometry('400x500')

# GUI layouts
label_EastTeams = tk.Label(master, text='East Teams', bg='yellow')
label_EastTeams.place(x=5, y=5)
label_WestTeams = tk.Label(master, text='West Teams', bg='cyan')
label_WestTeams.place(x=145, y=5)

# East Teams
val_Raptors = tk.IntVar()
checkbutton_Raptors = tk.Checkbutton(master, text='Toronto Raptors', variable=val_Raptors)
checkbutton_Raptors.place(x=5, y=25)
val_Bucks = tk.IntVar()
checkbutton_Bucks = tk.Checkbutton(master, text='Milwaukee Bucks', variable=val_Bucks)
checkbutton_Bucks.place(x=5, y=45)
val_76ers = tk.IntVar()
checkbutton_76ers = tk.Checkbutton(master, text='Philadelphia 76ers', variable=val_76ers)
checkbutton_76ers.place(x=5, y=65)
val_Pacers = tk.IntVar()
checkbutton_Pacers = tk.Checkbutton(master, text='Indiana Pacers', variable=val_Pacers)
checkbutton_Pacers.place(x=5, y=85)
val_Celtics = tk.IntVar()
checkbutton_Celtics = tk.Checkbutton(master, text='Boston Celtics', variable=val_Celtics)
checkbutton_Celtics.place(x=5, y=105)
val_Pistons = tk.IntVar()
checkbutton_Pistons = tk.Checkbutton(master, text='Detroit Pistons', variable=val_Pistons)
checkbutton_Pistons.place(x=5, y=125)
val_Hornets = tk.IntVar()
checkbutton_Hornets = tk.Checkbutton(master, text='Charlotte Hornets', variable=val_Hornets)
checkbutton_Hornets.place(x=5, y=145)
val_Magics = tk.IntVar()
checkbutton_Magics = tk.Checkbutton(master, text='Orlando Magics', variable=val_Magics)
checkbutton_Magics.place(x=5, y=165)
val_Heats = tk.IntVar()
checkbutton_Heats = tk.Checkbutton(master, text='Miami Heats', variable=val_Heats)
checkbutton_Heats.place(x=5, y=185)
val_Wizards = tk.IntVar()
checkbutton_Wizards = tk.Checkbutton(master, text='Washington Wizards', variable=val_Wizards)
checkbutton_Wizards.place(x=5, y=205)
val_Nets = tk.IntVar()
checkbutton_Nets = tk.Checkbutton(master, text='Brooklyn Nets', variable=val_Nets)
checkbutton_Nets.place(x=5, y=225)
val_Knicks = tk.IntVar()
checkbutton_Knicks = tk.Checkbutton(master, text='New York Knicks', variable=val_Knicks)
checkbutton_Knicks.place(x=5, y=245)
val_Hawks = tk.IntVar()
checkbutton_Hawks = tk.Checkbutton(master, text='Atlanta Hawks', variable=val_Hawks)
checkbutton_Hawks.place(x=5, y=265)
val_Cavaliers = tk.IntVar()
checkbutton_Cavaliers = tk.Checkbutton(master, text='Cleveland Cavaliers', variable=val_Cavaliers)
checkbutton_Cavaliers.place(x=5, y=285)
val_Bulls = tk.IntVar()
checkbutton_Bulls = tk.Checkbutton(master, text='Chicago Bulls', variable=val_Bulls)
checkbutton_Bulls.place(x=5, y=305)

# West Teams
val_Warriors = tk.IntVar()
checkbutton_Warriors = tk.Checkbutton(master, text='Golden State Warriors', variable=val_Warriors)
checkbutton_Warriors.place(x=145, y=25)
val_Thunders = tk.IntVar()
checkbutton_Thunders = tk.Checkbutton(master, text='Oklahoma City Thunders', variable=val_Thunders)
checkbutton_Thunders.place(x=145, y=45)
val_Nuggets = tk.IntVar()
checkbutton_Nuggets = tk.Checkbutton(master, text='Denver Nuggets', variable=val_Nuggets)
checkbutton_Nuggets.place(x=145, y=65)
val_Clippers = tk.IntVar()
checkbutton_Clippers = tk.Checkbutton(master, text='LA Clippers', variable=val_Clippers)
checkbutton_Clippers.place(x=145, y=85)
val_Lakers = tk.IntVar()
checkbutton_Lakers = tk.Checkbutton(master, text='Los Angeles Lakers', variable=val_Lakers)
checkbutton_Lakers.place(x=145, y=105)
val_Grizzlies = tk.IntVar()
checkbutton_Grizzlies = tk.Checkbutton(master, text='Memphis Grizzlies', variable=val_Grizzlies)
checkbutton_Grizzlies.place(x=145, y=125)
val_Blazers = tk.IntVar()
checkbutton_Blazers = tk.Checkbutton(master, text='Portland Trail Blazers', variable=val_Blazers)
checkbutton_Blazers.place(x=145, y=145)
val_Mavericks = tk.IntVar()
checkbutton_Mavericks = tk.Checkbutton(master, text='Dallas Mavericks', variable=val_Mavericks)
checkbutton_Mavericks.place(x=145, y=165)
val_Kings = tk.IntVar()
checkbutton_Kings = tk.Checkbutton(master, text='Sacramento Kings', variable=val_Kings)
checkbutton_Kings.place(x=145, y=185)
val_Timberwolves = tk.IntVar()
checkbutton_Timberwolves = tk.Checkbutton(master, text='Minnesota Timberwolves', variable=val_Timberwolves)
checkbutton_Timberwolves.place(x=145, y=205)
val_Pelicans = tk.IntVar()
checkbutton_Pelicans = tk.Checkbutton(master, text='New Orleans Pelicans', variable=val_Pelicans)
checkbutton_Pelicans.place(x=145, y=225)
val_Spurs = tk.IntVar()
checkbutton_Spurs = tk.Checkbutton(master, text='San Antonio Spurs', variable=val_Spurs)
checkbutton_Spurs.place(x=145, y=245)
val_Jazz= tk.IntVar()
checkbutton_Jazz = tk.Checkbutton(master, text='Utah Jazz', variable=val_Jazz)
checkbutton_Jazz.place(x=145, y=265)
val_Rockets = tk.IntVar()
checkbutton_Rockets = tk.Checkbutton(master, text='Houston Rockets', variable=val_Rockets)
checkbutton_Rockets.place(x=145, y=285)
val_Suns = tk.IntVar()
checkbutton_Suns = tk.Checkbutton(master, text='Phoenix Suns', variable=val_Suns)
checkbutton_Suns.place(x=145, y=305)



# Checkbuttons for column selection
label_Selection = tk.Label(master, text='Please Select ONLY 1 Statistic To Show!', bg='yellow')
label_Selection.place(x=5, y=330)
val_Rank = tk.IntVar()
checkbutton_Rank = tk.Checkbutton(master, text='Rank', variable=val_Rank)
checkbutton_Rank.place(x=5, y=350)
val_Pt_Gain = tk.IntVar()
checkbutton_Pt_Gain = tk.Checkbutton(master, text='Point Gain', variable=val_Pt_Gain)
checkbutton_Pt_Gain.place(x=65, y=350)
val_Pt_Loss = tk.IntVar()
checkbutton_Pt_Loss = tk.Checkbutton(master, text='Point Loss', variable=val_Pt_Loss)
checkbutton_Pt_Loss.place(x=145, y=350)
val_Pt_Diff = tk.IntVar()
checkbutton_Pt_Diff = tk.Checkbutton(master, text='Point Diff', variable=val_Pt_Diff)
checkbutton_Pt_Diff.place(x=225, y=350)
val_WinPctg = tk.IntVar()
checkbutton_WinPctg = tk.Checkbutton(master, text='Total Win %', variable=val_WinPctg)
checkbutton_WinPctg.place(x=5, y=370)
val_HomeWinPctg = tk.IntVar()
checkbutton_HomeWinPctg = tk.Checkbutton(master, text='Home Win %', variable=val_HomeWinPctg)
checkbutton_HomeWinPctg.place(x=95, y=370)
val_RoadWinPctg = tk.IntVar()
checkbutton_RoadWinPctg = tk.Checkbutton(master, text='Road Win %', variable=val_RoadWinPctg)
checkbutton_RoadWinPctg.place(x=195, y=370)



# Buttons
button_Update = tk.Button(master, text='Update!', command=function_Update)
button_Update.place(x=5, y=400)
button_Show = tk.Button(master, text='Show!', command=function_Plot)
button_Show.place(x=120, y=400)


master.mainloop()

